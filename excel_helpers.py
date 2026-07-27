"""
excel_helpers.py
-----------------
A tiny data-access layer that treats three .xlsx files as the database:

    admins.xlsx     -> AdminID, Name, Email, Phone, PasswordHash, CreatedAt
    customers.xlsx  -> CustomerID, Name, Email, Phone, PasswordHash,
                        TransactionHistory (JSON string), CreatedAt
    groceries.xlsx  -> ItemID, Name, Category, Unit, PricePerUnit,
                        QuantityInStock, ImageURL

Every function opens the workbook, does its work, saves, and closes again
inside the same call -- per the "quick tip" in the spec, keeping the file
locked for as short a time as possible. A process-wide lock serializes
access so two requests handled by the same worker can't corrupt a file by
writing at the same time.

This is intentionally simple and is meant for a single dev-server process.
It is NOT a substitute for a real database under concurrent multi-user
production traffic -- see the README for notes on that trade-off.
"""

import os
import json
from threading import Lock

import openpyxl
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Overridable so tests (and anyone who wants to) can point the app at an
# isolated data directory instead of the real one.
DATA_DIR = os.environ.get("GROCERY_APP_DATA_DIR", os.path.join(BASE_DIR, "data"))

ADMINS_FILE = os.path.join(DATA_DIR, "admins.xlsx")
CUSTOMERS_FILE = os.path.join(DATA_DIR, "customers.xlsx")
GROCERIES_FILE = os.path.join(DATA_DIR, "groceries.xlsx")

ADMIN_HEADERS = ["AdminID", "Name", "Email", "Phone", "PasswordHash", "CreatedAt"]
CUSTOMER_HEADERS = [
    "CustomerID", "Name", "Email", "Phone", "PasswordHash",
    "TransactionHistory", "CreatedAt",
]
GROCERY_HEADERS = [
    "ItemID", "Name", "Category", "Unit", "PricePerUnit",
    "QuantityInStock", "ImageURL",
]

# Guards every read/write below so concurrent requests in the same process
# can't interleave their file operations.
_lock = Lock()


# ---------------------------------------------------------------------------
# Initialization
# ---------------------------------------------------------------------------

def _init_file(filepath, headers):
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(headers)
        wb.save(filepath)


def init_excel_files():
    """Create the data directory and all three workbooks (with headers) if
    they don't already exist. Call this once when the app starts."""
    os.makedirs(DATA_DIR, exist_ok=True)
    _init_file(ADMINS_FILE, ADMIN_HEADERS)
    _init_file(CUSTOMERS_FILE, CUSTOMER_HEADERS)
    _init_file(GROCERIES_FILE, GROCERY_HEADERS)


# ---------------------------------------------------------------------------
# Reads
# ---------------------------------------------------------------------------

def _read_rows(filepath):
    with _lock:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if raw_row[0] in (None, ""):
                continue  # skip blank trailing rows
            rows.append(dict(zip(headers, raw_row)))
        wb.close()
        return rows


def get_all_rows(filepath):
    """Every record in the sheet, as a list of dicts keyed by header."""
    return _read_rows(filepath)


def get_row_by_id(filepath, id_column, id_value):
    """First record whose id_column matches id_value (compared as strings,
    so it doesn't matter whether you pass an int or a str), or None."""
    for row in _read_rows(filepath):
        if row.get(id_column) is not None and str(row[id_column]) == str(id_value):
            return row
    return None


def get_row_by_email(filepath, email):
    """Case-insensitive email lookup -- used for both login and the
    "no duplicate registrations" uniqueness check."""
    if not email:
        return None
    target = email.strip().lower()
    for row in _read_rows(filepath):
        stored = row.get("Email")
        if stored and str(stored).strip().lower() == target:
            return row
    return None


# ---------------------------------------------------------------------------
# Writes
# ---------------------------------------------------------------------------

def create_record(filepath, id_column, data):
    """Insert a new row with an auto-incremented ID (max existing ID + 1,
    or 1 for the first row). Returns the new ID. The ID is computed and the
    row appended inside a single locked open/save so two near-simultaneous
    signups can't be handed the same ID."""
    with _lock:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        id_idx = headers.index(id_column)

        existing_ids = [
            int(row[id_idx])
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[id_idx] is not None
        ]
        new_id = (max(existing_ids) + 1) if existing_ids else 1

        record = dict(data)
        record[id_column] = new_id
        ws.append([record.get(h, "") for h in headers])
        wb.save(filepath)
        wb.close()
        return new_id


def update_row(filepath, id_column, id_value, updates):
    """Update the first row whose id_column matches id_value with the
    key/value pairs in `updates`. Returns True if a row was found and
    updated, False otherwise."""
    with _lock:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if id_column not in headers:
            wb.close()
            return False

        id_idx = headers.index(id_column)
        found = False
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[id_idx].value is not None and str(row_cells[id_idx].value) == str(id_value):
                for key, value in updates.items():
                    if key in headers:
                        row_cells[headers.index(key)].value = value
                found = True
                break

        if found:
            wb.save(filepath)
        wb.close()
        return found


# ---------------------------------------------------------------------------
# Transaction history helpers (stored as a JSON string inside one cell)
# ---------------------------------------------------------------------------

def parse_transaction_history(raw):
    """Turn the TransactionHistory cell (a JSON string) into a Python list."""
    if not raw:
        return []
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def serialize_transaction_history(history_list):
    return json.dumps(history_list)
